#### Import lib ####

import pandas as pd
import openpyxl as op
import os as os
import re as re
from collections import Counter
import xlsxwriter as xl
from collections import deque

### Variable Declaration & Initialization ###

## Local disk file ##
old_directory_path = "C:\\Users\\user\\PycharmProjects\\ExcelProject\\Folder A"
new_directory_path = "C:\\Users\\user\\PycharmProjects\\ExcelProject\\Folder B"

## Create empty list to store data ##

# store file
old_directory_file = []
new_directory_file = []
c_directory_file = []

# store rfid
key_pattern_old_directory = []
key_pattern_new_directory = []
key_pattern_c_directory = []

# store data bucket location in worksheet
location_list = []
empty_row_list = []
data_values_rows = []
data_values_cols = []
empty_col_list = []
overall_empty_col_list = []
final_empty_col_list = []


## Counters ##
counter = 0
loc_counter = 0
df_counter = 1

## Pattern ##
file_key_pattern = "^\w{3}\d{3}"


## Function ##

# Remove duplicates within list #
def unique_list(list1,list2):
    for i in list1:
        if i not in list2:
            list2.append(i)
    return list2

# Append all files into directory #
def file_list(path,list_name):
    for file in os.listdir(path):
        filename = os.fsdecode(file)
        if filename.endswith(".xlsx"):
            list_name.append(file)
    return list_name

# Extract RFID key pattern from filename then save in list #
def find_rfid(rfid_pattern, list_name, save_to_list):
    for item in list_name:
        rfid_name = re.findall(rfid_pattern, item)
        for i in rfid_name:
            save_to_list.append(str(i))
    return save_to_list
######################################################################################################

## Loop through all files within Old Directory & Append to list ##
## Check file list ##
print("Old Direcotry: ",file_list(old_directory_path,old_directory_file))
print("Other Directory: ",file_list(new_directory_path,new_directory_file))


## Extract key pattern from workbook name ##
## ABC123, ABC124 >> detected as RFID ##
# find_rfid(file_key_pattern, old_directory_file,key_pattern_old_directory)
# find_rfid(file_key_pattern, c_directory_file, key_pattern_c_directory)

print("\n")

## Looking for matches in between file name, saved as list ##
matches_list = list({j:i for i,j in zip(old_directory_file,new_directory_file) if re.match(j,i)})


## Major loop ##

print("\nBegin")

for item in old_directory_file:
    if len(old_directory_file) >= 0:

        ## Creating file path to open workbook
        print(old_directory_file[counter])
        create_file_path = str(old_directory_path + "\\" + old_directory_file[counter])
        print("\nOpening Workbook = ", create_file_path)

        ## Start to open workbook
        wb = op.load_workbook(create_file_path)
        sheet_counter = 0

        ## start to iterate through worksheets, loop
        # for sheets in range(len(wb.sheetnames)):
        for sheets_counter in range(len(wb.sheetnames)):
            wb.active = sheet_counter

            if wb.active is not None:
                ## Wprk on active worksheet
                print("\nCurrent Active Sheet: ", wb.active)
                ws = wb.active


                ## Find all free entry locations in sheet
                for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
                    for cell in row:
                        if cell.value == 'Data Entry':
                            loc = cell.row
                            location_list.append(loc)
                ## Check item in list
                print("\nStarting point data bucket = ", location_list)

                ## Check all rows (entire) that are empty within worksheet, added restricted end_col
                end_col = 40
                for row in range(ws.max_row+1, 0,-1):
                    # print(row," ",[cell.value for cell in ws[row][0:ws.max_column]])
                    if all([cell.value is None for cell in ws[row][0:end_col+1]]):
                        if row > location_list[0]:
                            empty_row_list.append(row)
                            empty_row_list.sort(reverse=False)

                print("Empty row in worksheet: ",empty_row_list)

                ## Find first empty_col
                col_index = 0

                for col in ws.iter_cols(min_row=1, max_row=1, min_col= 1, max_col= 20):
                    for cell in col:
                        if cell.value is None:
                            empty_col_list.append(cell.column_letter)


                for col_index in range(len(empty_col_list)):
                    for col in ws[str(empty_col_list[col_index])]:
                        if all([col.value is None for col in ws[str(empty_col_list[col_index])]]):
                            overall_empty_col_list.append(col.col_idx)

                print("Empty Column in worksheet: ",unique_list(overall_empty_col_list,final_empty_col_list))


                ## Iniating counter for location_list[] and empty_row_list[], starting with 0 so that it iterates through the first element
                i = 0
                j = 0
                zero_index = final_empty_col_list[0]

                # ## Create a new workbook that will resides in C Directory, using Openpyxl as the ExcelWriter
                # test_name = 'ABC123_Val_Model_2021.xlsm'
                # temp_name = 'ABC123_Val_Model_2021.xlsx'
                # create_new_sheet_name = str(new_paste_directory_path + "\\" + test_name)
                # create_temp_sheet_name = str(new_paste_directory_path + "\\" + temp_name)
                # # new_workbook = op.load_workbook(create_new_sheet_name, keep_vba=True)
                # writer = pd.ExcelWriter(create_temp_sheet_name, engine='openpyxl')


                ## Outer loop that ensure that all data buckets list are iterated upon
                for i,j in zip(range(len(location_list)),range(len(empty_row_list))):
                    ## Ensure no data bucket is lesser than one another
                    if empty_row_list[j] < location_list[i]:
                       empty_row_list.pop(j)

                    ## The range of data bucket that is currently being worked on
                    print("\n", range(location_list[i], empty_row_list[j]))
                    ## Inner loop that focuses on extracting data into specified list, iterating through all rows, columns in order to get value, use dynamic data range based on row,col lists
                    data = [[ws.cell(r, c).value for c in range(1, zero_index)] for r in range(location_list[i], empty_row_list[j])]


                    ## Find the percentage of "None" (No value) from data ##
                    if sum(len(items) for items in data) != 0:
                        null_percentage = (sum(item.count(None) for item in data) / (sum(len(items) for items in data))*100.0)
                        # print("Percentage of No Value data: ",round(null_percentage,0), "%") ## to help developer understand
                    else:
                        continue


                    ## Converion to dataframe
                    if round(null_percentage,0) <= 50:
                        df = pd.DataFrame(data)

                        ## Reverse order based on dataframe index (to match data output in worksheet)
                        # df = df[df.columns[::-1]] #use this in case of inverted data values
                        print(df)
                        # df.to_excel(writer,sheet_name=str(wb.active), startrow=location_list[i], index_label=None, index=False,header=False)



                    else:
                        continue


                    i+=1
                    j+=1


                # writer.close()


                # Empty list to bring in new locations for another sheet
                location_list.clear()
                empty_row_list.clear()
                empty_col_list.clear()
                overall_empty_col_list.clear()
                final_empty_col_list.clear()


                # break
                sheet_counter+=1


        # Move to the next file
        counter += 1
        # break

        print("\n============================================================")